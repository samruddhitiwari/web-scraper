from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook

# Start a new instance of Chrome WebDriver
driver = webdriver.Chrome()

# Navigate to Google Docs
driver.get("https://docs.google.com/document/d/your_document_id/edit")

# Wait for the document to load
time.sleep(5)

# Find and extract the text content from the document
doc_content = driver.find_element(By.XPATH, "//div[@aria-label='Document content']")
text_content = doc_content.text.split('\n')  # Split the text content by newline to get rows

# Close the browser
driver.quit()

# Create a new Excel workbook
wb = Workbook()
ws = wb.active

current_heading = None  # Variable to keep track of the current heading

# Write the extracted text content into Excel rows and columns
row_idx = 1
for line in text_content:
    if line.startswith("Heading"):
        current_heading = line
    else:
        if current_heading:
            ws.cell(row=row_idx, column=1, value=current_heading)
            ws.cell(row=row_idx, column=2, value=line)
            row_idx += 1

# Save the workbook to a file
wb.save("extracted_data.xlsx")